import matplotlib
matplotlib.use('Agg')  # Kein GUI, nur Dateiausgabe


import pandas as pd
import logging
from . import config
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import styles
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages

logger = logging.getLogger(__name__)


def ws_customize(ws):

    for i, col in enumerate(ws.iter_cols(), start=1):

        col_letter = get_column_letter(i)

        col_width = 0
        for cell in col:

            if not cell.value == None:
                cell_len = len(str(cell.value))

                col_width = max(col_width, cell_len)

        ws.column_dimensions[col_letter].width = col_width + 2

    for row in range(1, ws.max_row + 1):
        ws.row_dimensions[row].height = 20

    thin = styles.Side(style="thin", color="000000")
    for col in ws:
        for cell in col:
            cell.alignment = styles.Alignment(horizontal="center", vertical="center")

            if cell.value in config.SPORTS.values():
                cell.fill = styles.PatternFill("solid", fgColor="000000")
            else:
                cell.fill = styles.PatternFill("solid", fgColor="4F81BD")
                cell.font = styles.Font(bold=True, color="FFFFFF")
                cell.border = styles.Border(
                    left=thin, right=thin, top=thin, bottom=thin
                )


def write_semestrs_title(ws, df: dict):
    title_line = 1

    for semestr, value in df.items():

        ws.merge_cells(f"A{title_line}:C{title_line}")
        ws.cell(row=title_line, column=1).value = f"{semestr}"
        ws.cell(row=title_line, column=1).fill = styles.PatternFill(
            "solid", fgColor="000000"
        )

        title_line += len(value) + 5

def max_cols_width(tabl):
           
    max_width = 0
            
    for key, cell in tabl.get_celld().items():

        text = cell.get_text().get_text()
        width = len(text) * 0.1
                    
        if not text or text == 'nan':
            cell.get_text().set_text('')
            continue
                    
        if max_width < width:
            max_width = width
            
    max_width = min(max_width, 0.5)
    
    for key, cell in tabl.get_celld().items():
        cell.set_width(max_width)


def df_toExel(df: dict):

    if config.FINALL_FILE_SAVE_AS == 'excel':
    
        start_row = 1

        with pd.ExcelWriter(config.OUTPUT_FILE, engine="openpyxl", mode="w") as writer:

            for value in df.values():

                logger.debug(f"Adding semestr at {start_row}:\n{value}")

                value.to_excel(writer, sheet_name="All", index=False, startrow=start_row)

                start_row += len(value) + 5

        wb = load_workbook(config.OUTPUT_FILE)
        ws = wb["All"]

        ws_customize(ws=ws)
        write_semestrs_title(ws=ws, df=df)

        wb.save(config.OUTPUT_FILE)
    
    elif config.FINALL_FILE_SAVE_AS == 'pdf':
        
        with PdfPages(r"Output\finall_tabel.pdf") as pdf:
            
            for semestr, sem_df in df.items():
                
                fig, ax = plt.subplots(figsize = (7, 8))
                
                ax.axis('off')
                
                tabl = plt.table(cellText= sem_df.values, colLabels= sem_df.columns, cellLoc= 'center', loc= 'center')
                
                tabl.scale(1.3, 1.3)
                tabl.auto_set_font_size(False)
                tabl.set_fontsize(10)
                
                plt.suptitle(f"{semestr}")
                
                max_cols_width(tabl)
                
                
                plt.tight_layout()
                pdf.savefig()
                plt.close()
            
            plt.close('all')
            
